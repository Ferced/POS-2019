#!/usr/bin/env python
# -*- coding: cp1252 -*-
# -*- coding: 850 -*-
# -*- coding: utf-8 -*-
import os
import random
import sqlite3
import ssl
import urllib
import re
import urllib2
import getpass
import requests
from bs4 import BeautifulSoup
from mechanize import Browser
import time
import datetime
from PIL import ImageTk, Image
#openpyxl
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from openpyxl.workbook import Workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from openpyxl.chart.series import DataPoint
def soloNum(documento):
    docuLimpio=""
    listaNum=["0","1","2","3","4","5","6","7","8","9"]
    for elemento in documento:
        if elemento in listaNum:
            docuLimpio+=elemento
    return docuLimpio
def listaDocus(larca):

    wolo3=1
    woloMD="A"
    colFinal=6
    AC1=True
    AC2=False
    AC3=False
    hojaLarca=larca.active
    if int(wolo3)>1 and int(wolo3)<50:
        hojaLarca.cell(row=int(wolo3)-1,column=int(colFinal)).value="PUCO"
        hojaLarca.cell(row=int(wolo3)-1,column=int(colFinal)+1).value="IOMA"
        hojaLarca.cell(row=int(wolo3)-1,column=int(colFinal)+2).value="SUPERINTENDENCIA"
        hojaLarca.cell(row=int(wolo3)-1,column=int(colFinal)+3).value="FECHA DE ALTA"
    hojaLarca.column_dimensions[hojaLarca.cell(row= 1, column=int(colFinal)).column].width=40
    hojaLarca.column_dimensions[hojaLarca.cell(row= 1, column=int(colFinal)+1).column].width=30
    hojaLarca.column_dimensions[hojaLarca.cell(row= 1, column=int(colFinal)+2).column].width=40
    hojaLarca.column_dimensions[hojaLarca.cell(row= 1, column=int(colFinal)+3).column].width=15
    n=1
    cadaTanto=1000
    contaNone=0
    lebrak=False
    default=50
    ejeYlista=default
    conta=0
    if AC1==True:
        for fila in hojaLarca:
            if contaNone>5:
                break
            for columna in fila:
                if n>=int(wolo3):
                    if columna.coordinate==woloMD.upper()+str(n):
                        if str(type(columna.value))!="<type 'NoneType'>":
                            try:
                                contaNone=0
                                pucoList2=str(botPUCO2018(soloNum(str(columna.value))))
                                print pucoList2
                                if len(pucoList2) > 200 and pucoList2 != "No se reportan datos":
                                    pucoList2="No se reportan datos"
                                hojaLarca.cell(row=n,column=int(colFinal)).value=pucoList2
                            except Exception as e:
                                pucoList2=str(botPUCO2018(soloNum(str(columna.value))))
                                if  len(pucoList2) > 20 and pucoList2 != "No se reportan datos":
                                    pucoList2="No se reportan datos"
                                hojaLarca.cell(row=n,column=int(colFinal)+1).value=pucoList2
                            ejeYlista+=20
                            if ejeYlista>580:
                                ejeYlista=default
                            print "-----------------------",n,"-------------------------"
                            print ejeYlista
                        else:
                            contaNone+=1
                        if contaNone>5:
                            break
                        if n%cadaTanto==0:
                            print "-----------------------------------------------------------------------"
                            print "GUARDE"
                            print "-----------------------------------------------------------------------"
                            # larca.save(rutaRelativa(nomDeArch)+nomFinal+"EnProceso1"+".xlsx")
                    conta=conta+1
            n+=1
        print "-----------------------------------TERMINO PUCO----------------------------------"
        print "-----------------------------------TERMINO PUCO----------------------------------"
        print "*********************************************************************************"
        # larca.save(rutaRelativa(nomDeArch)+nomFinal+".xlsx")
    if AC3 == True:
        n=1
        contaNone=0
        lebrak=False
        conta=0
        print "-----------------------------------EMPIEZA IOMA----------------------------------"
        print "-----------------------------------EMPIEZA IOMA----------------------------------"
        ejeYlista=default
        for fila in hojaLarca:
            if contaNone>5:
                break
            for columna in fila:
                if n>=int(wolo3):
                    if columna.coordinate==woloMD.upper()+str(n):
                        if str(type(columna.value))!="<type 'NoneType'>":
                            contaNone=0 
                            try:
                                if chequearIOMA(str(columna.value)) == True:
                                    hojaLarca.cell(row=n,column=int(colFinal)+1).value="TIENE IOMA"
                                else:
                                    hojaLarca.cell(row=n,column=int(colFinal)+1).value="NO TIENE IOMA"
                                ejeYlista+=20
                                if ejeYlista>580:
                                    ejeYlista=default
                            except Exception as e:
                                print "error de ioma"
                                print e
                                print "----------------------------------"
                                # time.sleep(61)
                                # hojaLarca.cell(row=n,column=int(colFinal)+1).value=str(chequearIOMA(soloNum(str(columna.value))))
                            print "-----------------------",n,"-------------------------"
                        else:
                            contaNone+=1
                        if contaNone>5:
                            break
                        if n%cadaTanto==0:
                            print "-----------------------------------------------------------------------"
                            print "GUARDE"
                            print "-----------------------------------------------------------------------"
                            # larca.save(rutaRelativa(nomDeArch)+nomFinal+"EnProceso2"+".xlsx")
                    conta=conta+1
            n+=1
        print "-----------------------------------TERMINO IOMA----------------------------------"
        print "-----------------------------------TERMINO IOMA----------------------------------"
        print "**********************************************************************************"
        # larca.save(rutaRelativa(nomDeArch)+nomFinal+".xlsx")
    if AC2 == True:
        n=1
        ejeYlista=default
        print "-----------------------------------EMPIEZA SUPERINTENDENCIA----------------------------------"
        print "-----------------------------------EMPIEZA SUPERINTENDENCIA----------------------------------"
        cantNone=0
        d=0
        for fila in hojaLarca:
            if cantNone>5:
                break
            for columna in fila:
                if n>=int(wolo3):
                    if columna.coordinate==woloMD.upper()+str(n):
                        if str(type(columna.value))!="<type 'NoneType'>" :
                            cantNone=0
                            try:
                                OSSI2=controlSPIN(soloNum(str(columna.value)))
                                ejeYlista+=20
                                if ejeYlista>580:
                                    ejeYlista=default
                                hojaLarca.cell(row=n,column=int(colFinal)+2).value=OSSI2
                                hojaLarca.cell(row=n,column=int(colFinal)+3).value=str(bateria)
                                print "-------------------------",n,"------------------------------"
                                if n%cadaTanto==0:
                                    print "GUARDE------------------------------------------"
                                    print "GUARDE------------------------------------------"
                                    print "GUARDE------------------------------------------"
                                    # larca.save(rutaRelativa(nomDeArch)+nomFinal+"EnProceso3"+".xlsx")
                            except Exception as e:
                                print "ERROR---------------------------------------"
                                print e
                                print "ERROR---------------------------------------"
                                try:
                                    time.sleep(61)
                                    OSSI2=controlSPIN(soloNum(str(columna.value)))
                                except Exception as b:
                                    print "ERROR2--------------------------------------"
                                    print e
                                    print "ERROR2--------------------------------------"
                        if str(type(columna.value))!="<type 'NoneType'>" :
                            cantNone+=1
                        if cantNone>5:
                            # larca.save(nomFinal+"EnProceso2.xlsx")
                            break
            n+=1
        print "-----------------------------------TERMINO SUPERINTENDENCIA----------------------------------"
        print "-----------------------------------TERMINO SUPERINTENDENCIA----------------------------------"
        print "**********************************************************************************"
    # larca.save(rutaRelativa(nomDeArch)+nomFinal+".xlsx")
    return larca
def limpiarNombre(nombre):
	acum=""
	ordenamiento=0
	for elemento in nombre:
		ordenamiento=ord(elemento)
		if ordenamiento >=65 and ordenamiento <=122 or ordenamiento==32:
			acum+=elemento
		else:
			if ordenamiento == 209:
				acum+="Ñ"
			else:
				acum+="-"
	return acum
def chequearSPINespecial(documento):
    print "USANDO SPIN ESPECIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAL----------------------------"
    global bateria
    global porta
    global nombreglobal2
    url = "https://seguro.sssalud.gob.ar/indexss.php?opc=bus650&user=HPGD&cat=consultas"
    nombreglobal2=""
    cadenita2 = ""
    cadenita = "" 
    bateria = ""
    franco = False
    encontro = False 
    encontro2 = False
    browser.open(url,timeout=2)
    form = browser.select_form(nr=0)
    browser["nro_doc"]=documento
    response =browser.submit()
    contadorsito2=0
    contadorsitocade2=0
    OSSI2=""
    leer = response.read()
    for elemento in leer:
        if encontro2 == True and elemento == "<":
            if contadorsitocade2 == 6:
                nombreglobal2 = cadenita2
            encontro2=False
            cadenita2=""
            contadorsitocade2+=1
        if encontro == True and elemento == "<":
            contadorsito2+=1
            if cadenita.strip(" ") == "No se reportan datos para el NUMERO DE DOCUMENTO "+documento:
                OSSI2="No se encontraron datos"
                return OSSI2
            if contadorsito2 == 3:
                OSSI2=cadenita
            if contadorsito2 == 4:
                bateria=cadenita
            encontro=False
            cadenita=""
        if elemento == "<":
            cadenita2=""
            cadenita=""
        cadenita2+=elemento
        cadenita+=elemento
        if cadenita2== '<td>':
            encontro2=True
            cadenita2=""
        if cadenita == '<b>':
            encontro=True
        if elemento==">":
            cadenita=""
    return OSSI2        
def controlSPIN(documento):
    try:
        ssl._create_default_https_context = ssl._create_unverified_context
        url = "https://seguro.sssalud.gob.ar/indexss.php?opc=bus650&user=HPGD&cat=consultas"
        header = {'User-Agent': 'Mozilla/5.0 (Windows NT 5.1; rv:14.0) Gecko/20100101 Firefox/14.0.1', 'Referer': 'http://whateveritis.com'}
        request = urllib2.Request(url, None, header)
        response = browser.open(request,timeout=2)
        form = browser.select_form(nr=0)
        try:
            browser["_user_name_"] =usuarioL
            browser["_pass_word_"] =contraL
            response =browser.submit()
            OSSI2=chequearSPIN(documento,response)
            return OSSI2
        except:
            OSSI2=chequearSPINespecial(documento)
            return OSSI2
    except Exception as e:
        print "----------------------------------------------------------------------------------------"
        print e
        print "----------------------------------------------------------------------------------------"
        return "ERROR"      
def chequearSPIN(documento,response):
    global bateria
    global porta
    global nombreglobal2
    nombreglobal2=""
    cadenita2 = ""
    cadenita = "" 
    bateria = ""
    franco = False
    encontro = False 
    encontro2 = False
    form = browser.select_form(nr=0)
    browser["nro_doc"]=documento
    response =browser.submit()
    contadorsito2=0
    contadorsitocade2=0
    OSSI2=""
    leer = response.read()
    for elemento in leer:
        if encontro2 == True and elemento == "<":
            if contadorsitocade2 == 6:
                nombreglobal2 = cadenita2
            encontro2=False
            cadenita2=""
            contadorsitocade2+=1
        if encontro == True and elemento == "<":
            contadorsito2+=1
            if cadenita.strip(" ") == "No se reportan datos para el NUMERO DE DOCUMENTO "+documento:
                OSSI2="No se encontraron datos"
                return OSSI2
            if contadorsito2 ==3:
                OSSI2=cadenita
            if contadorsito2 == 4:
                bateria=cadenita
            encontro=False
            cadenita=""
        if elemento == "<":
            cadenita2=""
            cadenita=""
        cadenita2+=elemento
        cadenita+=elemento
        if cadenita2== '<td>':
            encontro2=True
            cadenita2=""
        if cadenita == '<b>':
            encontro=True
        if elemento==">":
            cadenita=""
    return OSSI2
def chequearIOMA2018(documento):
	tiene=False
	for item in range(1,3):
		if chequearIOMA2018bis(documento,item) == True:
			return True
	return False
def chequearIOMA2018bis(documento,sexo):
	data = urllib.urlencode({'T3':documento,'sexo':sexo,'B13':'Buscar'})
	url="http://www.ioma.gba.gov.ar/sistemas/consulta_padron_afiliados/buscadorpordocumento.php"
	request = urllib2.Request(url,data)
	respuesta = urllib2.urlopen(request).read()
	counterIOMA=0
	itworks=False
	suma=0
	cadenita = ""
	ultima = ""
	franco = False
	encontro = False
	for elemento in respuesta:
		if encontro == True and elemento == "<":
			counterIOMA+=1
			if counterIOMA == 2:
				nombreglobal3 = cadenita
			ultima=cadenita
			encontro=False
			cadenita=""
		if elemento == "<":
			cadenita=""
		cadenita+=elemento
		if cadenita == '<span class="texto-azul">' or cadenita =='<span class="texto-azul-bold">' :
			encontro=True
			itworks = True
		if elemento==">":
			cadenita=""
	if len(ultima)>2:
		return True
	else:
		return False
def chequearIOMA(documento):
    global nombreglobal3
    nombreglobal3=""
    sexo=["1","2"]
    counterIOMA=0
    itworks=False
    suma=0
    try:
        for indele in sexo:
            cadenita = ""
            ultima = ""
            url = "http://www.ioma.gba.gov.ar/sistemas/consulta_padron_afiliados/consulta_afiliados.php" ##Son necesarias las comillas
            franco = False
            encontro = False
            browser.open(url)
            form = browser.select_form("numdoc")
            browser[" T3"] = documento
            browser.form.set_value([indele],name='sexo')
            response = browser.submit()
            leer = response.read()
            for elemento in leer:
                if encontro == True and elemento == "<":
                    counterIOMA+=1
                    if counterIOMA == 2:
                        nombreglobal3 = cadenita
                    ultima=cadenita
                    encontro=False
                    cadenita=""
                if elemento == "<":
                    cadenita=""
                cadenita+=elemento
                if cadenita == '<span class="texto-azul">' or cadenita =='<span class="texto-azul-bold">' :
                    encontro=True
                    itworks = True
                if elemento==">":
                    cadenita=""
            if len(ultima) > 2:
                suma+=1
            else:
                suma+=0
        if suma>0:
            return True
        else:
            return False
    except Exception as e:
            print "-----------------------------------------------------------------"
            print str(e)
            return "ERROR"
def botPUCOfalso(documento):
    return ["PAGINA NO DISPONIBLE"]
def botPUCO2(documento):
    global nombreglobal
    nombreglobal = ""
    data = urllib.urlencode({'documento':documento})
    try:
        url = "http://138.0.104.200/nacer/puco_historico.php"
        # url = "http://138.0.104.200/nacer/puco"
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        respLimpia,nombreglobal=limpiandingSTR(respuesta)
        if respuesta == "null":
            return "No se reportan datos"
        else:
            return respLimpia
    except: 
        listaEr=["PAGINA NO DISPONIBLE"]
        return listaEr
def limpiandingSTR(cadena):
    nombresito=""
    c=""
    nombre=""
    listreturn=[]
    conta=1
    conta2=0
    eselnombre=False
    encontro = False
    termino= False
    nueva=0
    for elemento in cadena:
        if eselnombre == True:
            if elemento == '"':
                nombresito=nombre
                eselnombre=False
        nombre+=elemento
        if elemento == ",":
            nombre=""
        if elemento != '"' and elemento != "}" and encontro == True: 
            c+=elemento
        if elemento == ":":
            c=""
            encontro =True
        if elemento == "}":
            if c not in listreturn and len(c)>3:
                listreturn.append(c)
            conta2=0
        if nombre == '"NombreYApellido":"':
            eselnombre=True
            nombre=""
    return listreturn,nombresito
def limpiandingSTR2(cadena):
    global estaactivosumar
    clausula = False
    clausula2 = False
    clausula3= False
    apellido=""
    nombreactivo=""
    eselapellido=False
    nombresito=""
    nombresito2=""
    c=""
    nombre=""
    listreturn=[]
    conta=1
    conta2=0
    eselnombre=False
    encontro = False
    termino= False
    nueva=0
    for elemento in cadena:
        if eselnombre == True:
            if elemento == '"' and clausula == False:
                nombresito=nombre
                eselnombre=False
        if eselapellido == True:
            if elemento == '"' and clausula2 == False:
                nombresito2=apellido
                eselapellido=False
        apellido+=elemento
        nombre+=elemento
        nombreactivo+=elemento
        if elemento == "," and encontro == True:
            nombre=""
        if elemento != '"' and elemento != "}" and encontro == True: 
            c+=elemento
        if elemento == ":":
            c=""
            encontro =True
        if elemento == "}":
            listreturn.append(c)
            conta2=0
        if nombre == '"afiNombre":"':
            eselnombre=True
            nombre=""
        if nombre == '"afiApellido":"':
            eselapellido=True
            apellido=""
        # if nombre == '"Activo":"N"' and clausula3 == False:
        #   clausula3= True
        #   estaactivosumar=False
        if nombre == '"Activo":"S"' and clausula3 == False:
            estaactivosumar = True
            return nombresito+" "+nombresito2
    estaactivosumar = False
    return ""
def chequearPuco(documento):
    try:
        listaOS=[]
        noTiene=False
        url = "http://www.saludnqn.gob.ar/PadronConsultasWeb/"
        browser.open(url,timeout=2)
        form = browser.select_form(nr=0)
        browser["ctl00$ContentPlaceHolder1$txtNumero"]=documento
        response =browser.submit()
        respuesta=response.read()
        counter=0
        notieneOS=False
        oeses=[5,11,17,23,29,35]
        cadenita = ""
        ultima = ""
        franco = False
        encontro = False
        for elemento in respuesta:
                if encontro == True and elemento == "<":
                    counter+=1
                    ultima=cadenita
                    if ultima.strip(" ") == "No se encontraron datos":
                        notieneOS=True
                        return noTiene
                    encontro=False
                    cadenita=""
                if elemento == "<":
                    cadenita=""
                cadenita+=elemento
                if cadenita == '<td colspan="6">':
                    encontro=True
                if elemento==">":
                    cadenita=""
        cadenita=""
        counter=0
        ultima=""
        encontro=False
        if notieneOS==False:
            for elemento in respuesta:
                if encontro == True and elemento == "<":
                    counter+=1
                    ultima=cadenita
                    if counter in oeses:
                        listaOS.append(ultima.strip(" "))
                    encontro=False
                    cadenita=""
                if elemento == "<":
                    cadenita=""
                cadenita+=elemento
                if cadenita == '<td style="font-size:Smaller;">' or cadenita == '<div class="text">':
                    encontro=True
                if elemento==">":
                    cadenita=""
        return listaOS
    except Exception as e:
            print str(e)
            return "ERROR"
def conseguirInput(cadena):
    inputLimpio=""
    corchetes=0
    encontroInput=False
    pasoLos2Car=False
    conta=0
    for elemento in cadena:
        if encontroInput==True and conta == 2 and elemento=="'":
            return inputLimpio
        if elemento == "[" and encontroInput==False:
            corchetes+=1
        if encontroInput==True and conta<2:
            conta+=1
        if corchetes==2:
            encontroInput=True
        if conta==2:
            if elemento!="'":
                inputLimpio+=elemento
def printAlgo():
    print"*********************************************************************************************"
    print "holaaaaaaaaaaaaaaaaaaaa"
    print"*********************************************************************************************"
def siLaPalabraEstaEnLaLista(lista,palabra):
    for elemento in lista:
        if palabra in elemento:
            return True
    return False
def tienePAMI(laStr, elemento1, elemento2):
    if laStr in elemento1 or siLaPalabraEstaEnLaLista(elemento2,laStr):
        return True
    return False
def chequearPAMI(documento):
    browser.open('http://institucional.pami.org.ar/result.php?c=6-2')
    form=browser.select_form(nr=3)
    browser.set_handle_robots( False )
    browser["nroDocumento"]=documento
    response =browser.submit()
    respuesta=response.read()
    return filtrarVomitoV2(respuesta)
def filtrarVomito(vomito):
    url=""
    acum=""
    acumLink=""
    encontroLink=False
    loproximo=False
    lodeahora=False
    for elemento in vomito:
        if lodeahora==True: 
            if elemento !='"':
                acumLink+=elemento
            else:
                if "beneficio" in acumLink and "parent" in acumLink and encontroLink == False:
                    encontroLink=True
                    url=acumLink
                lodeahora=False
                acumLink=""
        if elemento == "<":
            if loproximo==True:
                loproximo=False
        acum+=elemento
        if elemento == ">":
            if acum == '<p class="whitetxt">':
                loproximo=True
            acum=""
        if acum=='<a href="':
            lodeahora=True
    url="http://institucional.pami.org.ar/"+url
    return chequearArbolPami(url)
def filtrarVomitoV2(vomito):
    global contaRAROS
    listaLinks=[]
    listaNom=[]
    listaAlta=[]
    listaBaja=[]
    listaDoc=[]
    listadoPami=[]
    url=""
    acum=""
    acumLink=""
    encontroLink=False
    loproximo=False
    lodeahora=False
    contaRAROS=0
    lista=["td"]
    cunta=0
    soup = BeautifulSoup(vomito,'html.parser')
    for elemento in soup.findAll('p', attrs={'class': 'whitetxt'}):
        cunta+=1
        # print elemento.text
        if cunta==1:
            listaNom.append(elemento.text)
        if cunta==4:
            listaDoc.append(elemento.text)
        if cunta==5:
            listaAlta.append(elemento.text)
        if cunta%6==0:
            listaBaja.append(elemento.text)
            # raw_input()
            cunta=0
    contaLetras=""
    for link in soup.findAll('a'):
        if link.string=="add":
            for elemento in str(link):
                contaLetras+=elemento 
                if elemento == '"':
                    if "href" in contaLetras:
                        contaLetras=""
                    if "beneficio" in contaLetras and "parent" in contaLetras:
                        listaLinks.append(contaLetras.rstrip('"').replace("amp;", ''))
                        contaLetras=""
    cuntaLE=0
    listado1=[]
    listado2=[]
    listado3=[]
    for elemento in listaLinks:
        if len(listaBaja[cuntaLE]) < 2:
            print listaNom[cuntaLE]
            print listaDoc[cuntaLE]
            print listaAlta[cuntaLE]
            print listaBaja[cuntaLE]
            print listaLinks[cuntaLE]
            url="http://institucional.pami.org.ar/"+listaLinks[cuntaLE]
            listado1,listado2,listado3= chequearArbolPami(url)
            listadoPami.append([listaDoc[cuntaLE],listaNom[cuntaLE]])
            listadoPami.append(zip(listado1,listado2,listado3))


        cuntaLE+=1
    cuntaLE=0
    # for lista in listadoPami:
    #     if cuntaLE % 2 == 0:
    #         for a,b,c in lista:
    #             print a,"  ",b,"   ",c
    #     else:
    #         for a in lista:
    #             print a
    #     cuntaLE+=1
    return listadoPami
def chequearArbolPami(url):
    r  = requests.get(url)
    listaDmodulo=[]
    listaRed=[]
    listaPrestador=[]
    contador=0
    contaTabla=1
    data = r.text
    soup = BeautifulSoup(data,"html5lib")
    for link in soup.find_all('p'):
        # print(link.get('class'))
        if "<p>PRESTADOR:</p>" in str(link): 
            contador+=1
        if contador==2 and "<p>PRESTADOR:</p>" not in str(link):
            linkLimpio=str(link).strip("<p>").strip("</p>")
            if "APELLIDOS Y NOMBRES" in str(link):
                break
            if contaTabla==4:
                contaTabla=1
            if contaTabla==1:
                # print "D.MODULO: "+linkLimpio
                listaDmodulo.append(str(linkLimpio))
            if contaTabla==2:
                # print "RED: "+linkLimpio
                listaRed.append(str(linkLimpio))
            if contaTabla==3:
                # print "PRESTADOR: "+linkLimpio
                listaPrestador.append(str(linkLimpio))
            contaTabla+=1
    return listaDmodulo,listaRed,listaPrestador
def nombreLigadoAlDocumento():
    if len(nombreglobal)>4:
        return nombreglobal
    if len(nombreglobal3)>4:
        return nombreglobal3
    if len(nombreglobal2)>4:
        return nombreglobal2
    return False
def tieneOS(IOMA,PUCO,SPIN):
	tiene=False
	if IOMA != False and IOMA != "ERROR":
		return True
	if SPIN != "No se encontraron datos" and SPIN != "ERROR":
		return True 
	if PUCO != ["No se reportan datos"] and len(PUCO[0]) > 1:
		return True
	return tiene
def limpiarNombres():
	global nombreglobalsumar
	global nombreglobal
	global nombreglobal2
	global nombreglobal3
	nombreglobalsumar=""
	nombreglobal=""
	nombreglobal2=""
	nombreglobal3=""
def botPUCO2018(documento):
    global nombreglobal
    nombreglobal = ""
    data = urllib.urlencode({'documento':documento})
    try:
        url = 'http://138.0.104.200/nacer/personas_puco_con_documentos.php'
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        respLimpia,nombreglobal=limpiandingSTR(respuesta)
        if respuesta == "null":
            return "No se reportan datos"
        else:
            return respLimpia
    except: 
        return "ERROR"
def botPAMI2018(documento):
    data = urllib.urlencode({'tipoDocumento':'DNI','nroDocumento':documento,'submit2':'Buscar'})
    url='http://institucional.pami.org.ar/result.php?c=6-2-2'
    # request = urllib2.Request(url,data)
    # respuesta = urllib2.urlopen(request).read()
    respuesta = requests.post(url,data=data)
    # filtrarVomito(respuesta)
    print respuesta.text
    print filtrarVomito(respuesta)
    return listaDmodulo,listaRed,listaPrestador
def botPUCOhistorico(documento):
    data = urllib.urlencode({'nrodoc':documento,"tabla":"PUCO_2018-05"})
    url = "http://138.0.104.200/nacer/puco.php"
    request = urllib2.Request(url,data) 
    respuesta = urllib2.urlopen(request).read()
    respLimpia=limpiandingSTR(respuesta)
    if respuesta == "null":
        return False
    else:
        return respLimpia

def escribirEnLaBD(documento,nombre,puco,ioma,spin):
   puco2=""
   nombre=nombre.strip("- ")
   spin=spin.strip(" ")
   if type(puco) == list:
       for elemento in puco:
           puco2+=elemento+", "
       puco=puco2
   conn = sqlite3.connect(os.path.join( os.path.dirname( __file__),'..')+"/static/pos.sqlite3") 
   cursor = conn.cursor()
   conn.text_factory = str
   cursor.execute('''INSERT INTO historial(documento,nombre,puco,ioma,spin,fecha) VALUES(:documento, :nombre, :puco, :ioma, :spin, :fecha)''',                  {'documento':documento, 'nombre':nombre, 'puco':puco, 'ioma':ioma, 'spin':spin, 'fecha':datetime.datetime.now() })
   conn.commit()
   cursor.execute('SELECT * FROM historial')
   data = cursor.fetchall()

def autoTestingPOS(documento):
    url="http://172.20.40.88:81/POS/consultaUnica/"
    browser.open(url)
    form = browser.select_form("installer")
    browser["your_name"] = documento
    response = browser.submit()
    leer = response.read()
    listaParametros=["p","strong"]
    print entreParametros(listaParametros,leer)

def sacarEstosElementos(cadena,lista):
    for elemento in lista:
        cadena=cadena.strip(elemento)
    return cadena
# sacarEstosElemntos("</strong>",["<","/",">"])
def botPUCO2018(documento):
    global intentoListaDoc
    global nombreglobal
    nombreglobal = ""
    data = urllib.urlencode({'documento':documento})
    try:
        url = "http://138.0.104.200/nacer/personas_puco_con_documentos.php"
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        respLimpia,nombreglobal=limpiandingSTR(respuesta)
        if respuesta == "null" or len(respLimpia) == 0:
            return "No se reportan datos"
        else:
            return respLimpia
    except:
        listaEr=["PAGINA NO DISPONIBLE"]
        return listaEr
def botSPIN2018(documento):
    global intentoListaDoc
    global nombreglobal
    payload = {'_user_name_':'sss1754','_pass_word_':'1w4un8'}
	# url = 'https://seguro.sssalud.gob.ar/login.php?opc=bus650&user=HPGD&cat=consultas'
	# try:
	# 	print requests.post(url,header,data=payload)
	# except Exception as e:
	# 	request=urllib2.Request(url,payload,header)
	# 	respuesta = urllib2.urlopen(request).read()
	# 	print respuesta
    ssl._create_default_https_context = ssl._create_unverified_context
    header = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/67.0.3396.99 Safari/537.36'}
    url = 'https://seguro.sssalud.gob.ar/login.php?opc=bus650&user=HPGD&cat=consultas'
    login_data = dict(login='sss1754', password='1w4un8')
    session = requests.session()
    # session.headers.update(header)
    # urllib3.contrib.pyopenssl.inject_into_urllib3()
    r = session.post(url, data=payload, headers=header, verify='cacert.pem')
    print r.text #prints the <html> response.
    r2 = session.get('https://seguro.sssalud.gob.ar/login.php?opc=bus650&user=HPGD&cat=consultas')
    print r2.content #prints the raw html you can now parse and scrape
    # r = requests.post(url, auth=('sss1754','1w4un8'), verify = True)
    # print r
    # raw_input()
    # data = urllib.urlencode({'q': 'Status'})

    # h = httplib.HTTPConnection('myserver:8080')

    # headers = {"Content-type": "application/x-www-form-urlencoded", "Accept": "text/plain"}

    # h.request('POST', '/inout-tracker/index.php', data, headers)

    # r = h.getresponse()

    # print r.read()
    # nombreglobal = ""
    # ssl._create_default_https_context = ssl._create_unverified_context
    # header = {'User-Agent': 'Mozilla/5.0 (Windows NT 5.1; rv:14.0) Gecko/20100101 Firefox/14.0.1', 'Referer': 'http://whateveritis.com'}
    # header = {'_user_name_':'sss1754','_pass_word_':'1w4un8','submitbtn':'ingresar'}
    # header = {'User-Agent': 'Mozilla/5.0 (Windows NT 5.1; rv:14.0) Gecko/20100101 Firefox/14.0.1', 'Referer': 'http://whateveritis.com'}
    # respuesta = requests.get('https://seguro.sssalud.gob.ar/login.php?opc=bus650&user=HPGD&cat=consultas', auth=('sss1754', '1w4un8'))
    # data = urllib.urlencode({'nro_doc':documento,'B1':'Conjhhjasdsa'})
    # # url = "https://seguro.sssalud.gob.ar/login.php?opc=bus650&user=HPGD&cat=consultas"
    # url = "https://seguro.sssalud.gob.ar/indexss.php?opc=bus650&user=HPGD&cat=consultas"
    # request = urllib2.Request(url,data,header)
    # respuesta = urllib2.urlopen(request).read()
    # print respuesta
def botPUCO2018lista(listaDoc,listaReal):
    global intentoListaDoc
    global nombreglobal
    listaDocumentos=[]
    listaObrasSociales=[]
    listaNombres=[]
    par=0
    listaBan=[",",":"]
    nombreglobal = ""
    data = urllib.urlencode({'documento':listaDoc})
    try:
        print "ENTROOOOOOOOOOOOOOOOOOOOOOOOOOOOO"
        print listaDoc
        print "---------------------------"
        print listaReal
        url = "http://138.0.104.200/nacer/personas_puco_con_documentos.php"
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        regex = r"\{(.*?)\"}"
        elDocumento=""
        esElDoc=False
        ultimo=""
        elNombre=""
        matches = re.finditer(regex, respuesta, re.MULTILINE | re.DOTALL)
        for matchNum, match in enumerate(matches):
            par=0
            for groupNum in range(0, len(match.groups())):       
                elmo=match.group(1)
                esElDoc=False
                for elemento in elmo.split('"'):
                    if elemento not in listaBan:
                        # print elemento
                        if par%2==0:
                            if par==4:
                                elDocumento=elemento
                            if par==6:
                                elNombre=elemento
                            # print "--------------------------------------------------------------"
                        ultimo=elemento

                        par+=1
                # print "******************************************"
                listaDocumentos.append(elDocumento)
                listaNombres.append(elNombre)
                listaObrasSociales.append(ultimo)
        print listaDocumentos
        lezipeado=zip(listaDocumentos,listaNombres,listaObrasSociales)
        return lezipeado
    except Exception as e:
        print e 
        listaEr=["PAGINA NO DISPONIBLE"]
        return listaEr
    return listaDocumentos,listaNombres,listaObrasSociales

def nombreSumar():
	global nombreglobalsumar
	return (nombreglobalsumar)
def botPUCOSUMAR(documento):
    try:
        global nombreglobalsumar
        data = urllib.urlencode({'documento':documento})
        url = "http://138.0.104.200/nacer/personas_inscriptas_hist_con_documentos.php"
        request = urllib2.Request(url,data)
        respuesta = urllib2.urlopen(request).read()
        nombreglobalsumar=limpiandingSTR2(respuesta)
        if respuesta == "null" or estaactivosumar == False:
            return False
        else:
            return True
    except:
        return "ERROR"
def dosListasv2(lista1A,lista2B,listaOS,listaNom):
    acum=0
    for elemento in lista2B:
        print elemento
        print listaOS[acum]
        print listaNom[acum]
        acum+=1
        print "-------------------------------------------------"
    lista3C=[]
    lista3Cos=[]
    lista3Nom=[]
    acum=0
    acumEspecial=0
    anterior=""
    for elemento in lista2B:
        while lista1A[acum] != elemento:
            lista3C.append(lista1A[acum])
            lista3Cos.append("NO")
            lista3Nom.append("NO")
            if acum<=len(lista1A)-1:
                acum+=1
                break
        lista3C.append(lista1A[acum])
        lista3Cos.append(listaOS[acumEspecial])
        lista3Nom.append(listaNom[acumEspecial])
        anterior = elemento
        acumEspecial+=1
    while acum<=len(lista1A)-1:
        lista3C.append(lista1A[acum])
        lista3Cos.append("NO")
        lista3Nom.append("NO")
        acum+=1
    acum=0
    anterior=""
    while acum <= len(lista3C)-1:
        if anterior == lista3C[acum] and lista3Cos[acum]=="NO":
            if lista3Cos[acum-1]!="NO":
                lista3Cos.pop(acum)
                lista3C.pop(acum)
                lista3Nom.pop(acum)
                acum-=1
        anterior=lista3C[acum]
        acum+=1
    acum=0
    return convertirEnUna(zip(lista3C,lista3Cos,lista3Nom))

def entreParametros(parametros,texto):
    lista=["<","/",">"]
    abrio1=False
    cerro1=False
    abrio2=False
    denovo=False
    TextoEtiquetas=[]
    acumulador=""
    textoNec=""
    abrioDV=True
    cerroDV=False
    parametroNecesario=True
    # for elemento in texto:
    #     if elemento == ">":
    #         abrio2=True
    #         print acumulador
    #     if elemento == "<":
    #         if abrio1==True and abrio2 == True:
    #             abrio2=False
    #             acumulador=""
    #         if abrio1==True and abrio2==False:
    #             print textoNec
    #         abrio1=True
    #     if abrio1==True or abrio2==True:
    #         if elemento not in lista:
    #             acumulador+=elemento
    #     if abrio1==True and abrio2==False:
    #         textoNec+=elemento
    for elemento in texto:
        if elemento == "<":
            if parametroNecesario==True:
                if len(textoNec) >1:
                    if "DATOS" not in textoNec:
                        TextoEtiquetas.append(textoNec.strip(" ")) 
                parametroNecesario=False
            textoNec=""
            denovo=False
            abrio1=True
        if denovo==True:
            abrio1=False
            cerro1=False
        if abrio1==True and cerro1==False:
            acumulador+=elemento
        if denovo ==True:
            if abrioDV==True:
                textoNec+=elemento
        if elemento == ">":
            cerro1=True
            conn = sqlite3.connect(os.path.join( os.path.dirname( __file__),'..')+"/pos.sqlite3") 
            # print acumulador
            # print sacarEstosElementos(acumulador,lista)
            if sacarEstosElementos(acumulador,lista) in parametros:
                parametroNecesario=True
            if "/" in acumulador:
                cerroDV,abrioDV=True,False
            else:
                abrioDV,cerroDV=True,False
            acumulador=""
            denovo=True
    return TextoEtiquetas
def convertirEnUna(masiva):
    listaDOC=[]
    listaOS=[]
    listaNOM=[]
    anterior=0
    acum=0
    for doc,os,nom in masiva:
        if anterior==doc:
            listaOS[acum-1]=listaOS[acum-1]+" ; "+os
        else:
            listaDOC.append(doc)
            listaOS.append(os)
            listaNOM.append(nom)
            acum+=1
        anterior=doc
    return zip(listaDOC,listaOS,listaNOM)
def listaDocus2018bis(hojaDocus):
    columnaDocus="A"
    n=1
    listaDocus=[]
    contaNones=0
    for fila in hojaDocus:
        if contaNones>9:
                break
        for columna in fila:
            if contaNones>9:
                break
            if columna.coordinate==columnaDocus+str(n):
                if columna.value==None or len(str(columna.value))<7 and columna.value>1000000:
                    contaNones+=1
                else:
                    contaNones=0
                    listaDocus.append(str(columna.value))
        n+=1
    return listaDocus
def botPUCO2018listaBis(lista):
    listaLimpia=[]
    acum=0
    listaDoc=[]
    listaNom=[]
    listaOSES=[]
    listaDoc1=[]
    listaNom1=[]
    listaOSES1=[]
    registro1=""
    registro2=""
    nidea=""
    listadocus=""
    for elemento in lista:
        listadocus+=filtroSoloNumeros(elemento)+" "
        if acum%2500==0 and acum>1 or acum+1==len(lista):
            niidea=botPUCO2018lista(listadocus.strip(" "),lista)
            for a,b,c in niidea:
                listaDoc.append(a)
                listaNom.append(b)
                listaOSES.append(c)
            listadocus=""
        acum+=1
    acum=0
    acum2=0
    # return dosListasv2(lista,listaDoc,listaOSES,listaNom)
    return filtroDeListongui(listaDoc,listaOSES,listaNom)
def filtroSoloNumeros(numero):
    acum=""
    if numero != None:
        for elemento in numero:
            if elemento.isdigit() == True:
                acum+=elemento
        return acum
def listaDocus2018(archivo):
    hojaDocumentos=archivo.active
    listaTerminada=botPUCO2018listaBis(listaDocus2018bis(hojaDocumentos))
    print type(listaTerminada)
    vacio=Workbook()
    hojaDocumentos=vacio.active
    n=1
    hojaDocumentos.column_dimensions["C"].width=75
    hojaDocumentos.column_dimensions["B"].width=35
    for a,b,c in listaTerminada:
        if c == "NO":
            hojaDocumentos["A"+str(n)].value=a
            hojaDocumentos["B"+str(n)].value="NO HAY DATOS"
            hojaDocumentos["C"+str(n)].value="NO HAY DATOS"
        else:
            hojaDocumentos["A"+str(n)].value=a
            hojaDocumentos["B"+str(n)].value=c
            hojaDocumentos["C"+str(n)].value=b
        n+=1
    return vacio
def filtroDeListongui(listaDocus,listaOses,listaNom):
    conta=0
    anterior=""
    todo=zip(listaDocus,listaOses,listaNom) 
    while conta<=len(listaDocus)-1:
        if anterior==listaDocus[conta]:
            listaOses[conta-1]=listaOses[conta-1]+" ; "+listaOses[conta]
            listaOses.pop(conta)
            listaDocus.pop(conta)
            listaNom.pop(conta)
            conta=conta-1
        anterior=listaDocus[conta]
        conta+=1
    return zip(listaDocus,listaOses,listaNom)
def chequearPAMIconTIPO(documento):
    listaOpciones=["DNI","LC","LE","PAS","CI"]
    # listaOpciones=["DNI"]
    vomitoLimpio=[]
    listaZips=[]
    listaTipo=[]
    for elemento in listaOpciones:
        print elemento
        browser.open('http://institucional.pami.org.ar/result.php?c=6-2')
        form=browser.select_form(nr=3)
        browser.set_handle_robots( False )
        browser["nroDocumento"]=documento
        browser["tipoDocumento"]=[elemento]
        response =browser.submit()
        respuesta=response.read()
        lista1a,lista2b,lista3b=filtrarVomito(respuesta)
        vomitoLimpio=zip(lista1a,lista2b,lista3b)
        if len(vomitoLimpio)>1 and len(vomitoLimpio[0][0])>0:
            listaZips.append(elemento)
            listaZips.append(vomitoLimpio)

    acum=0
    return listaZips
def masInformacionBots(documento):
    listaZips=[]
    listaTipos=[]
    sumar=""
    listaZips=chequearPAMIconTIPO(documento)
    sumar=botPUCOSUMAR(documento)
    return listaZips,sumar

#MAS TE VALE NO ROMPER NADA HIJO DE PUTA------------------
global intentoListaDoc
global browser
global usuarioL
global contraL
global nombreglobal
global nombreglobal2
global nombreglobal3
global nombreglobalsumar
listaDoc=[]
listaNom=[]
listaOSES=[]
nombreglobal=""
nombreglobal2=""
nombreglobal3=""
usuarioL="sss1754"
contraL="1w4un8"
browser=Browser()
browser.set_handle_robots(False)

# for elemento in listao4:
#     print elemento
# print "-----------------------------------------------------------------------------"
# listaRandoms=[]
# for i in range(10):
#     listaRandoms.append(str(random.randint(10000,100000000)))
# for elemento in listaRandoms:
#     print elemento
# for a,b,c in botPUCO2018listaBis(listaRandoms):
#     print a,"  ",b,"  ",c
#     print " "
# listaRandoms=[]
# while True:
#     print chequearPAMI(str(random.randint(10000,100000000)))

# lista=["style","p"]
# texto="sdaasffhajshljakjfklasjfklajflkj<style>bueeenas</style>iasdjoasjfoiadasfsajfkajskfljasfjl<p>comooooooooooooooooooremavamos</p>jasoifjsaofjsaoifjoais"
# print len(chequearPAMI("11822431"))
# print chequearPAMI("6828448")
# listaPamisNormal_=chequearPAMI("11822431")
# print siLaPalabraEstaEnLaLista(listaPamisNormal_[0][2],"LARCADE")